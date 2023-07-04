import sys
import json
from PyQt5 import QtWidgets, QtGui, QtCore
from PyQt5.QtWidgets import QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton, QTableWidget, QTableWidgetItem, QComboBox, QMessageBox, QTabWidget, QGridLayout, QSpinBox, QStyleFactory, QCheckBox, QDialog, QInputDialog
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont, QPalette, QColor,QBrush
from collections import defaultdict
import datetime
import portalocker
import getpass
import openpyxl
from PyQt5.QtCore import QTimer
from PyQt5.QtGui import QKeySequence
from PyQt5.QtWidgets import QMainWindow, QShortcut, QDialog, QVBoxLayout, QLabel, QComboBox, QPushButton, QHBoxLayout, QApplication, QListWidget, QListWidgetItem


INVENTORY_FILE = "inventory.json"
CHECKEDOUT_FILE = "checked_out.json"
ACP_FILE = "acp.json"

# Laden der Modelle aus der JSON-Datei
with open('models.json', 'r') as f:
    MODELS = json.load(f)

# Beispiel zum Laden der ACP-Daten aus der JSON-Datei
def load_acp(self):
    try:
        with open(ACP_FILE, 'r') as f:
            self.acp = json.load(f)
    except FileNotFoundError:
        self.acp = []    

LANGUAGES = ["DE", "UK"]    

class SerialNumberDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Seriennummer ändern")
        self.setModal(True)

        layout = QVBoxLayout(self)

        self.serial_number_label = QLabel("Neue Seriennummer:")
        layout.addWidget(self.serial_number_label)

        self.serial_number_edit = QLineEdit()
        layout.addWidget(self.serial_number_edit)

        buttons_layout = QHBoxLayout()
        layout.addLayout(buttons_layout)

        self.accept_button = QPushButton("OK")
        self.accept_button.clicked.connect(self.accept)
        buttons_layout.addWidget(self.accept_button)

        self.cancel_button = QPushButton("Abbrechen")
        self.cancel_button.clicked.connect(self.reject)
        buttons_layout.addWidget(self.cancel_button)

    def get_serial_number(self):
        return self.serial_number_edit.text()
    
class MoveDevicesDialog(QDialog):
    def __init__(self, inventory, acp, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Geräte verschieben")
        self.setModal(True)

        layout = QVBoxLayout(self)

        self.source_label = QLabel("Von:")
        layout.addWidget(self.source_label)

        self.source_combo_box = QComboBox()
        self.source_combo_box.addItem("Aktuelles Inventar", inventory)
        self.source_combo_box.addItem("ACP", acp)
        layout.addWidget(self.source_combo_box)

        self.destination_label = QLabel("Zu:")
        layout.addWidget(self.destination_label)

        self.destination_combo_box = QComboBox()
        self.destination_combo_box.addItem("ACP", acp)
        self.destination_combo_box.addItem("Aktuelles Inventar", inventory)
        layout.addWidget(self.destination_combo_box)

        self.devices_label = QLabel("Geräte:")
        layout.addWidget(self.devices_label)

        self.devices_list = QListWidget()
        self.devices_list.setSelectionMode(QListWidget.MultiSelection)
        layout.addWidget(self.devices_list)

        buttons_layout = QHBoxLayout()
        layout.addLayout(buttons_layout)

        self.ok_button = QPushButton("OK")
        self.ok_button.clicked.connect(self.accept)
        buttons_layout.addWidget(self.ok_button)

        self.cancel_button = QPushButton("Abbrechen")
        self.cancel_button.clicked.connect(self.reject)
        buttons_layout.addWidget(self.cancel_button)

        self.source_combo_box.currentIndexChanged.connect(self.load_devices)

    def load_devices(self):
        selected_source = self.source_combo_box.currentData()
        self.devices_list.clear()

        if selected_source:
            for device in selected_source:
                item = QListWidgetItem(device['SN'])
                self.devices_list.addItem(item)

    def get_source(self):
        return self.source_combo_box.currentText()

    def get_destination(self):
        return self.destination_combo_box.currentText()

    def get_selected_devices(self):
        selected_devices = []
        for index in range(self.devices_list.count()):
            item = self.devices_list.item(index)
            selected_devices.append(item.text())
        return selected_devices




class MainWindow(QMainWindow):
    MODELS = ["HP Elitebook 835 G8", "HP Elitebook 645 G9", "HP Elitebook 655 G9","HP Elitebook X360 1040 G9","Dell Precision 5570","Dell Precision 7670","Dell Precision 7770", "Dell Optiplex 5000","Dell Optiplex 5000 Micro", "Dell Optiplex 7010 Tower", "Dell Precision T3660", "Dell Precision T7865 GOLD", "Dell Precison T7865 SILVER"]
    LANGUAGES = ["DE", "UK"]


    def on_mac_address_changed(self):
        mac_address = self.checkin_mac.text()
        mac_address = mac_address.replace(':', '')
        self.checkin_mac.setText(mac_address)
        
    def move_devices(self, source_list, destination_list, selected_items):
        for device in selected_items:
            if device in source_list:
                source_list.remove(device)
                destination_list.append(device)
        self.update_inventory()
        if destination_list is self.acp:
            self.update_acp()


    def open_move_devices_dialog(self):
        dialog = MoveDevicesDialog(self.inventory, self.acp)
        if dialog.exec_() == QDialog.Accepted:
            source = dialog.get_source()
            destination = dialog.get_destination()
            selected_devices = dialog.get_selected_devices()

            # Verschieben Sie die ausgewählten Geräte von der Quelle zum Ziel
            if source == "Aktuelles Inventar" and destination == "ACP":
                self.move_devices(self.inventory, self.acp, selected_devices)
            elif source == "ACP" and destination == "Aktuelles Inventar":
                self.move_devices(self.acp, self.inventory, selected_devices)

            self.save_acp()  # Speichern Sie die aktualisierten ACP-Daten
            self.save_inventory()  # Speichern Sie das aktualisierte Inventar
            self.update_inventory()  # Aktualisieren Sie die Bestandsanzeige



    def update_acp(self):
        acp_table = self.acp_table
        acp_table.setRowCount(0)  # Zurücksetzen der Zeilen in der Tabelle

        for item in self.acp:
            row_position = acp_table.rowCount()
            acp_table.insertRow(row_position)

            acp_table.setItem(row_position, 0, QTableWidgetItem(item.get('Modell', '-')))
            acp_table.setItem(row_position, 1, QTableWidgetItem(item.get('Sprache', '-')))
            acp_table.setItem(row_position, 2, QTableWidgetItem(item.get('SN', '-')))
            acp_table.setItem(row_position, 3, QTableWidgetItem(item.get('MAC-Adresse', '-')))
            acp_table.setItem(row_position, 4, QTableWidgetItem(item.get('Einbuchungsdatum', '-')))
            acp_table.setItem(row_position, 5, QTableWidgetItem(item.get('Ausbuchungsdatum', '-')))
            acp_table.setItem(row_position, 6, QTableWidgetItem(item.get('User', '-')))

    

    def __init__(self):
        super().__init__()
        self.MODELS = MODELS
        self.LANGUAGES = LANGUAGES
        self.checked_out = []
        self.initUI()
        self.acp = []

        # Erstellen eines Timers, um die Benutzeroberfläche regelmäßig zu aktualisieren
        self.update_timer = QTimer(self)
        self.update_timer.timeout.connect(self.load_data)
        self.update_timer.start(2000)  # Aktualisieren der Daten alle 1 Sekunden (1000 ms)

        self.checkin_mac.textChanged.connect(self.on_mac_address_changed)

        # Definieren der Toolbar
        self.toolbar = self.addToolBar("Main Toolbar")

        # Hinzufügen von Schaltflächen zur Symbolleiste
        self.toolbar.addWidget(QLabel("Suche: "))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Suche im Inventar...")
        self.toolbar.addWidget(self.search_input)

        # Erstellen Sie den Info-Button und fügen Sie ihn zur Symbolleiste hinzu
        self.info_button = QPushButton("Info")
        self.toolbar.addWidget(self.info_button)

        # Verbinden Sie den Button mit der Funktion show_info
        self.info_button.clicked.connect(self.show_info)

        # Verbinden der Suchfunktion mit der Suchschaltfläche
        self.search_input.textChanged.connect(self.on_search_text_changed)
        self.search_input.textChanged.connect(self.search_all)

        # Erstellen eines Shortcut für die gewünschte Tastenkombination (Strg+3)
        self.shortcut_ctrl_3 = QShortcut(QKeySequence("Ctrl+3"), self)
        self.shortcut_ctrl_3.activated.connect(self.show_dialog)

        self.load_acp()
        self.update_acp()

    def save_acp(self):
        with open(ACP_FILE, 'w') as f:
            json.dump(self.acp, f)

    def load_acp(self):
        try:
            with open(ACP_FILE, 'r') as f:
                self.acp = json.load(f)
        except FileNotFoundError:
            self.acp = []

    def show_dialog(self):
        dialog = MoveDevicesDialog(self.inventory, self.acp)  # Übergeben Sie das Inventar und die ACP-Daten an den Dialog
        if dialog.exec_() == QDialog.Accepted:
            source = dialog.get_source()
            destination = dialog.get_destination()
            selected_devices = dialog.get_selected_devices()

            # Verschieben der ausgewählten Geräte von der Quelle zum Ziel
            if source == "Aktuelles Inventar" and destination == "ACP":
                self.move_devices(self.inventory, self.acp, selected_devices)
            elif source == "ACP" and destination == "Aktuelles Inventar":
                self.move_devices(self.acp, self.inventory, selected_devices)

            self.save_acp()  # Speichern der aktualisierten ACP-Daten
            self.save_inventory()  # Speichern des aktualisierten Inventars
            self.update_inventory()  # Aktualisieren der Bestandsanzeige



    def show_info(self):
        # Erstellen Sie ein Informationsdialogfeld
        info_dialog = QMessageBox(self)
        info_dialog.setIcon(QMessageBox.Information)
        info_dialog.setWindowTitle("Information")
        info_dialog.setTextFormat(Qt.RichText)
        info_dialog.setText("<b>STRG + 1</b>: Änderungen in der Inventurliste vornehmen<br><b>STRG + 2</b>: Modell hinzufügen/umbenennen/entfernen")
        info_dialog.exec_()    


    def search_all(self):
        self.search_inventory()
        self.search_checkedout()
    

    def on_search_text_changed(self):
        # Überprüfen, ob das Suchfeld leer ist
        if not self.search_input.text().strip():
            # Wenn das Suchfeld leer ist, starten Sie den Timer
            self.update_timer.start()
        else:
            # Wenn das Suchfeld nicht leer ist, pausieren Sie den Timer
            self.update_timer.stop()    



    def search_inventory(self):
        query = self.search_input.text().lower().strip()

        if not query:
            # Wenn die Suchanfrage leer ist, zeige das gesamte Inventar
            self.update_inventory()
            # Starte den Timer, um das Inventar regelmäßig zu aktualisieren
            self.update_timer.start(2000)
            return

        filtered_inventory = []

        # Stoppe den Timer, um das Inventar nicht zu aktualisieren
        self.update_timer.stop()

        # Durchsuche das Inventar nach der Suchanfrage
        for item in self.inventory:
            if (query in item.get('Modell', '').lower()) or (query in item.get('SN', '').lower()):
                filtered_inventory.append(item)

        # Zeige die gefilterten Ergebnisse in der Tabelle an
        self.inventory_table.setRowCount(0)
        for item in filtered_inventory:
            row_position = self.inventory_table.rowCount()
            self.inventory_table.insertRow(row_position)

            self.inventory_table.setItem(row_position, 0, QTableWidgetItem(item['Modell']))
            self.inventory_table.setItem(row_position, 1, QTableWidgetItem(item['Sprache']))
            self.inventory_table.setItem(row_position, 2, QTableWidgetItem(item['SN']))
            self.inventory_table.setItem(row_position, 3, QTableWidgetItem(item.get('MAC-Adresse', '-')))
            self.inventory_table.setItem(row_position, 4, QTableWidgetItem(item.get('Einbuchungsdatum', '-')))
            self.inventory_table.setItem(row_position, 5, QTableWidgetItem(item.get('User', '-')))
  
    def closeEvent(self, event):
        self.search_input.clear()
        self.update_timer.stop()
        event.accept()

        

    def search_checkedout(self):
        # Erhalten der Suchanfrage aus dem Textfeld
        search_query = self.search_input.text().strip().lower()

        if not search_query:
            # Wenn die Suchanfrage leer ist, zeige das gesamte ausgecheckte Inventar
            self.update_checkedout()
            return

        # Erstellen einer neuen Liste mit den gefilterten Ergebnissen
        filtered_checkedout = [item for item in self.checked_out if search_query in str(item).lower()]

        # Leeren der Tabelle
        self.checkedout_table.setRowCount(0)

        # Einfügen der gefilterten Ergebnisse in die Tabelle
        for item in filtered_checkedout:
            row_position = self.checkedout_table.rowCount()
            self.checkedout_table.insertRow(row_position)

            self.checkedout_table.setItem(row_position, 0, QTableWidgetItem(item['Modell']))
            self.checkedout_table.setItem(row_position, 1, QTableWidgetItem(item['Sprache']))
            self.checkedout_table.setItem(row_position, 2, QTableWidgetItem(item['SN']))
            self.checkedout_table.setItem(row_position, 3, QTableWidgetItem(item.get('MAC-Adresse', '-')))
            self.checkedout_table.setItem(row_position, 4, QTableWidgetItem(item.get('Einbuchungsdatum', '-')))
            self.checkedout_table.setItem(row_position, 5, QTableWidgetItem(item.get('User', '-')))



    def initUI(self):
        # Haupt-Widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # Vertikales Layout für alle anderen Widgets
        layout = QVBoxLayout()
        central_widget.setLayout(layout)

        # Tab-Widget
        self.tab_widget = QTabWidget()
        layout.addWidget(self.tab_widget)

        # Tab für aktuelles Inventar
        self.inventory_tab = QWidget()
        self.tab_widget.addTab(self.inventory_tab, "Aktuelles Inventar")
        inventory_layout = QVBoxLayout()
        self.inventory_tab.setLayout(inventory_layout)

        # Container-Widget für die Tabellen
        table_container = QWidget()

        # Horizontales Layout für die Tabellen
        table_layout = QHBoxLayout(table_container)

        # Tabelle für die Anzeige des Bestands
        self.inventory_table = QTableWidget()
        self.inventory_table.setColumnCount(6)
        self.inventory_table.setHorizontalHeaderLabels(["Modell", "Sprache", "SN", "MAC-Adresse", "Einbuchungsdatum","User"])
        self.inventory_table.setColumnWidth(0, 250)
        self.inventory_table.setColumnWidth(1, 250)
        self.inventory_table.setColumnWidth(2, 215)
        self.inventory_table.setColumnWidth(3, 250)
        self.inventory_table.setColumnWidth(4, 200)
        table_layout.addWidget(self.inventory_table)

        # Tabelle für die Anzeige des Aktuellen Bestands
        self.current_table2 = QTableWidget()
        self.current_table2.setColumnCount(2)
        self.current_table2.setHorizontalHeaderLabels(["", "Aktueller Bestand"])
        self.current_table2.setColumnWidth(0, 50)
        self.current_table2.setColumnWidth(1, 450)
        self.current_table2.setFixedSize(500, 1275) # Größe auf 500x200 Pixel festlegen
        table_layout.addWidget(self.current_table2, 1)

        # Anzahl der Einträge in der Tabelle "inventory_table" zählen
        row_count = self.inventory_table.rowCount()

        # Tabelle "current_table2" aktualisieren
        self.current_table2.setRowCount(1)
        self.current_table2.setItem(0, 0, QTableWidgetItem(str(row_count)))

        # Füge das Layout des Container-Widgets zum Haupt-Layout hinzu
        inventory_layout.addWidget(table_container)

        # Widget für das Einbuchen eines Notebooks
        self.checkin_widget = QWidget()
        checkin_layout = QHBoxLayout()
        self.checkin_widget.setLayout(checkin_layout)
        layout.addWidget(self.checkin_widget)

        # Labels und Comboboxen für das Einbuchen
        checkin_layout.addWidget(QLabel("Modell:"))
        self.checkin_model = QComboBox()
        self.checkin_model.addItems(self.MODELS)
        checkin_layout.addWidget(self.checkin_model)

        checkin_layout.addWidget(QLabel("Sprache:"))
        self.checkin_language = QComboBox()
        self.checkin_language.addItems(self.LANGUAGES)
        checkin_layout.addWidget(self.checkin_language)

        checkin_layout.addWidget(QLabel("SN:"))
        self.checkin_sn = QLineEdit()
        checkin_layout.addWidget(self.checkin_sn)

        # Widget für das Ausbuchen eines Notebooks
        self.checkout_widget = QWidget()
        checkout_layout = QHBoxLayout()
        self.checkout_widget.setLayout(checkout_layout)
        layout.addWidget(self.checkout_widget)

        # Label und Textfeld für das Ausbuchen
        checkout_layout.addWidget(QLabel("SN:"))
        self.checkout_sn = QLineEdit()
        checkout_layout.addWidget(self.checkout_sn)
        
        #Mac
        checkin_layout.addWidget(QLabel("MAC-Adresse:"))
        self.checkin_mac = QLineEdit()
        checkin_layout.addWidget(self.checkin_mac)

        # Button zum Einbuchen
        checkin_button = QPushButton("Einbuchen")
        checkin_button.clicked.connect(self.checkin)
        checkin_layout.addWidget(checkin_button)

        # Signal zum Einbuchen auslösen, wenn Enter gedrückt wird
        self.checkin_sn.returnPressed.connect(self.focus_mac_address)
        
        # Signal zum Bestätigen des Einbuchens auslösen, wenn Enter im MAC-Adressfeld gedrückt wird
        self.checkin_mac.returnPressed.connect(checkin_button.click)

        # Button zum Ausbuchen
        checkout_button = QPushButton("Ausbuchen")
        checkout_button.clicked.connect(self.checkout)
        checkout_button.clicked.connect(self.clear_checkout_sn)
        checkout_layout.addWidget(checkout_button)
        
        # Tab für ausgebuchte Notebooks
        self.checkedout_tab = QWidget()
        self.tab_widget.addTab(self.checkedout_tab, "Ausgebuchte Notebooks")
        checkedout_layout = QVBoxLayout()
        self.checkedout_tab.setLayout(checkedout_layout)

        # Tabelle für die Anzeige der ausgebuchten Notebooks
        self.checkedout_table = QTableWidget()
        self.checkedout_table.setColumnCount(7)
        self.checkedout_table.setHorizontalHeaderLabels(["Modell", "Sprache", "SN", "MAC-Adresse", "Einbuchungsdatum", "Ausbuchungsdatum", "User"])
        checkedout_layout.addWidget(self.checkedout_table)
        
        self.checkedout_table.setColumnWidth(0, 300)
        self.checkedout_table.setColumnWidth(1, 250)
        self.checkedout_table.setColumnWidth(2, 300)
        self.checkedout_table.setColumnWidth(3, 300)
        self.checkedout_table.setColumnWidth(4, 300)
        self.checkedout_table.setColumnWidth(5, 260)

        # Erstellen Sie eine neue Registerkarte für ACP
        self.acp_tab = QWidget()
        self.tab_widget.addTab(self.acp_tab, "ACP")
        acp_layout = QVBoxLayout()
        self.acp_tab.setLayout(acp_layout)

        # Erstellen Sie eine Tabelle für die ACP-Liste
        self.acp_table = QTableWidget()
        self.acp_table.setColumnCount(7)  # Ändern Sie diese Zahl entsprechend der Anzahl der benötigten Spalten
        self.acp_table.setHorizontalHeaderLabels(["Modell", "Sprache", "SN", "MAC-Adresse", "Einbuchungsdatum", "Ausbuchungsdatum", "User"])  # Ändern Sie diese Liste entsprechend Ihren Spalten
        acp_layout.addWidget(self.acp_table)

        self.acp_table.setColumnWidth(0, 300)
        self.acp_table.setColumnWidth(1, 250)
        self.acp_table.setColumnWidth(2, 300)
        self.acp_table.setColumnWidth(3, 300)
        self.acp_table.setColumnWidth(4, 300)
        self.acp_table.setColumnWidth(5, 260)



        # Fügen Sie den folgenden Code in die initUI-Methode ein, um den Export-Button zu erstellen:
        export_button = QPushButton("Exportieren zu Excel")
        export_button.clicked.connect(self.export_to_excel)
        layout.addWidget(export_button)

        # Verbindung des Signals returnPressed mit der checkout-Methode
        self.checkout_sn.returnPressed.connect(self.checkout)
        self.inventory = []

        self.inventory = [{"Modell": "", "Sprache": "", "SN": "", "Einbuchungsdatum": ""}]

        # Laden der Inventardaten aus der Datei
        try:
            with open(INVENTORY_FILE, 'r') as f:
                self.inventory = json.load(f)
        except FileNotFoundError:
                self.inventory = []

        # Laden der Inventardaten aus der Datei
        try:
            with open(INVENTORY_FILE, 'r') as f:
                self.inventory = json.load(f)
        except FileNotFoundError:
            self.inventory = []
            
        # Laden der ausgebuchten Notebooks aus der Datei
        try:
            with open(CHECKEDOUT_FILE, 'r') as f:
                self.checked_out = json.load(f)
        except FileNotFoundError:
            self.checked_out = []

        # Anzeigen des aktuellen Bestands
        self.update_inventory()
        self.update_checkedout()

        # Fenstereinstellungen
        self.setWindowTitle("Inventar-Tool")
        self.resize(1200, 1500)

        pass

    def load_data(self):
        with open(INVENTORY_FILE, 'r') as f:
            portalocker.lock(f, portalocker.LOCK_SH)
            self.inventory = json.load(f)
            portalocker.unlock(f)

        with open(CHECKEDOUT_FILE, 'r') as f:
            portalocker.lock(f, portalocker.LOCK_SH)
            self.checked_out = json.load(f)
            portalocker.unlock(f)

    def search_inventory(self):
        search_query = self.search_input.text().strip().lower()

        # Leeren der Tabelle
        self.inventory_table.setRowCount(0)

        # Einfügen der gefilterten Inventardaten in die Tabelle
        for item in self.inventory:
            # Überprüfen, ob der Suchbegriff in einem der Felder des Inventarobjekts vorkommt
            if any(search_query in str(value).lower() for value in item.values()):
                row_position = self.inventory_table.rowCount()
                self.inventory_table.insertRow(row_position)

                self.inventory_table.setItem(row_position, 0, QTableWidgetItem(item['Modell']))
                self.inventory_table.setItem(row_position, 1, QTableWidgetItem(item['Sprache']))
                self.inventory_table.setItem(row_position, 2, QTableWidgetItem(item['SN']))
                self.inventory_table.setItem(row_position, 3, QTableWidgetItem(item.get('MAC-Adresse', '-')))
                self.inventory_table.setItem(row_position, 4, QTableWidgetItem(item.get('Einbuchungsdatum', '-')))
                self.inventory_table.setItem(row_position, 5, QTableWidgetItem(item.get('User', '-')))

                # Zentrieren aller Einträge in der Tabelle
                for i in range(self.inventory_table.columnCount()):
                    item = self.inventory_table.item(row_position, i)
                    if item is not None:
                        item.setTextAlignment(Qt.AlignCenter)
        

    def update_ui(self):

        self.load_data
        # Aktualisieren der Bestandsanzeige und der ausgebuchten Notebooks
        self.update_inventory()
        self.update_checkedout()
    
    def show_error_message(self, message):
        error_dialog = QMessageBox()
        error_dialog.setIcon(QMessageBox.Critical)
        error_dialog.setText("Fehler")
        error_dialog.setInformativeText(message)
        error_dialog.setWindowTitle("Fehler")
        error_dialog.exec_()


        
    # Funktion zum Einbuchen eines Notebooks
    def checkin(self):
        self.load_data()
        user = getpass.getuser()
        modell = self.checkin_model.currentText()
        if not modell:
            return
                
        sprache = self.checkin_language.currentText()
        if not sprache:
            return
                
        sn_list = []
        while True:
            sn = self.checkin_sn.text()
            if not sn:
                break
            sn_list.append(sn)
            self.checkin_sn.clear()
            
        mac = self.checkin_mac.text()
        if not mac:
            self.show_error_message("Bitte geben Sie eine MAC-Adresse ein.")
            return

        now = datetime.datetime.now()
        
        for sn in sn_list:
        # Überprüfen, ob die SN bereits im Inventar vorhanden ist
            if any(d['SN'] == sn for d in self.inventory):
                error_msg = f"Die Seriennummer {sn} ist bereits im Bestand vorhanden."
                self.show_error_message(error_msg)
                continue

        # Hinzufügen des Notebooks zum Bestand
            self.inventory.append({
                "Modell": modell,
                "Sprache": sprache,
                "SN": sn,
                "MAC-Adresse": mac,  # Füge die MAC-Adresse zum Inventar hinzu
                "Einbuchungsdatum": now.strftime("%Y-%m-%d %H:%M:%S"),
                "User": user
            })

        # Speichern der Inventardaten in der Datei
        with open(INVENTORY_FILE, 'w') as f:
            json.dump(self.inventory, f)

        # Speichern der Inventardaten in der Datei
        self.save_inventory()
        # Aktualisieren der Bestandsanzeige
        self.update_inventory()

        # Leeren des Textfelds für die Seriennummer
        self.checkin_sn.clear()
            
        # Leeren des Textfelds für die MAC-Adresse
        self.checkin_mac.clear()
        # Setzen des Fokus auf das Textfeld für die Seriennummer
        self.checkin_sn.setFocus()

        self.save_inventory()
        self.load_data()  # Aktualisieren der Daten nach der Aktion
      
    # Funktion zum Ausbuchen eines Notebooks
    def checkout(self):
        self.load_data()
        sn = self.checkout_sn.text()
        if not sn:
            return
            
        now = datetime.datetime.now()

        
        # Überprüfen, ob die SN im Inventar vorhanden ist
        for item in self.inventory:
            if item['SN'] == sn:
                self.inventory.remove(item)
                item['Ausbuchungsdatum'] = now.strftime("%Y-%m-%d %H:%M:%S")
                item['User'] = getpass.getuser()
                self.checked_out.append(item)  # Füge das Notebook zur Liste der ausgebuchten Notebooks hinzu
                break
        else:
            # Überprüfen, ob die SN bereits ausgebucht wurde
            for item in self.checked_out:
                if item['SN'] == sn:
                    error_msg = f"Die Seriennummer {sn} wurde bereits am {item['Ausbuchungsdatum']} von {item['User']} ausgebucht."
                    self.show_error_message(error_msg)
                    break
            else:
                error_msg = f"Die Seriennummer {sn} ist nicht im Bestand vorhanden."
                self.show_error_message(error_msg)
                return

        # Speichern der Inventardaten in der Datei
        self.save_inventory()
        self.save_checkedout()

        # Aktualisieren der Bestandsanzeige und der ausgebuchten Notebooks
        self.update_inventory()
        self.update_checkedout()

        # Leeren des Textfelds
        self.checkout_sn.clear()
        self.checkout_sn.setFocus()

        self.save_inventory()
        self.load_data()  # Aktualisieren der Daten nach der Aktion
        
    def export_to_excel(self):
        # Erstellen einer neuen Arbeitsmappe
        wb = openpyxl.Workbook()

        # Hinzufügen von Arbeitsblättern für das Inventar und die ausgebuchten Notebooks
        inventory_ws = wb.active
        inventory_ws.title = "Inventar"
        checkedout_ws = wb.create_sheet("Ausgebuchte Notebooks")

        # Schreiben der Überschriften für das Inventar-Arbeitsblatt
        inventory_headers = ["Modell", "Sprache", "SN", "MAC-Adresse", "Einbuchungsdatum", "User"]
        for col_num, header in enumerate(inventory_headers, 1):
            cell = inventory_ws.cell(row=1, column=col_num)
            cell.value = header

        # Schreiben der Daten in das Inventar-Arbeitsblatt
        for row_num, item in enumerate(self.inventory, 2):
            for col_num, header in enumerate(inventory_headers, 1):
                cell = inventory_ws.cell(row=row_num, column=col_num)
                cell.value = item.get(header, "")

        # Schreiben der Überschriften für das ausgebuchte Notebook-Arbeitsblatt
        checkedout_headers = ["Modell", "Sprache", "SN", "MAC-Adresse", "Einbuchungsdatum", "Ausbuchungsdatum", "User"]
        for col_num, header in enumerate(checkedout_headers, 1):
            cell = checkedout_ws.cell(row=1, column=col_num)
            cell.value = header

        # Schreiben der Daten in das ausgebuchte Notebook-Arbeitsblatt
        for row_num, item in enumerate(self.checked_out, 2):
            for col_num, header in enumerate(checkedout_headers, 1):
                cell = checkedout_ws.cell(row=row_num, column=col_num)
                cell.value = item.get(header, "")

        # Speichern der Arbeitsmappe
        filename = f"Inventar_und_Ausgebuchte_Notebooks_{datetime.datetime.now().strftime('%Y_%m_%d')}.xlsx"
        wb.save(filename)

        # Anzeigen einer Meldung, dass der Export erfolgreich war
        msg_box = QMessageBox()
        msg_box.setWindowTitle("Export")
        msg_box.setText(f"Die Daten wurden erfolgreich in die Datei {filename} exportiert.")
        msg_box.exec_()
 
    def focus_mac_address(self):
        self.checkin_mac.setFocus()

    def clear_checkout_sn(self):
        # Leeren der Tabelle
        self.checkedout_table.setRowCount(0)
        user = getpass.getuser()
        # Einfügen der ausgebuchten Notebooks in die Tabelle
        for item in self.checked_out:
            row_position = self.checkedout_table.rowCount()
            self.checkedout_table.insertRow(row_position)

            self.checkedout_table.setItem(row_position, 0, QTableWidgetItem(item['Modell']))
            self.checkedout_table.setItem(row_position, 1, QTableWidgetItem(item['Sprache']))
            self.checkedout_table.setItem(row_position, 2, QTableWidgetItem(item['SN']))
            self.checkedout_table.setItem(row_position, 3, QTableWidgetItem(item['MAC-Adresse']))
            self.checkedout_table.setItem(row_position, 4, QTableWidgetItem(item.get('Einbuchungsdatum', '-')))
            self.checkedout_table.setItem(row_position, 5, QTableWidgetItem(item.get('User', '-')))

    def update_inventory(self):
        # Leeren der Tabelle
        self.inventory_table.setRowCount(0)

        # Einfügen der aktuellen Inventardaten in die Tabelle
        for item in self.inventory:
            row_position = self.inventory_table.rowCount()
            self.inventory_table.insertRow(row_position)

            self.inventory_table.setItem(row_position, 0, QTableWidgetItem(item['Modell']))
            self.inventory_table.setItem(row_position, 1, QTableWidgetItem(item['Sprache']))
            self.inventory_table.setItem(row_position, 2, QTableWidgetItem(item['SN']))
            self.inventory_table.setItem(row_position, 3, QTableWidgetItem(item.get('MAC-Adresse', '-')))
            self.inventory_table.setItem(row_position, 4, QTableWidgetItem(item.get('Einbuchungsdatum', '-')))
            self.inventory_table.setItem(row_position, 5, QTableWidgetItem(item.get('User', '-')))

        # Zentrieren aller Einträge in der Tabelle
        for i in range(self.inventory_table.rowCount()):
            for j in range(self.inventory_table.columnCount()):
                item = self.inventory_table.item(i, j)
                if item is not None:
                    item.setTextAlignment(Qt.AlignCenter)

            
        #Erstelle ein Dictionary, um die Anzahl jedes Modells zu zählen
        model_counts = defaultdict(int)
        for item in self.inventory:
            model_counts[(item['Modell'], item['Sprache'])] += 1

        # Füge die Daten in die Tabelle ein
        self.current_table2.setRowCount(len(model_counts))
        for i, (model, count) in enumerate(model_counts.items()):
            count_item = QTableWidgetItem(str(count))
            count_item.setTextAlignment(Qt.AlignCenter)
            model_item = QTableWidgetItem(f"{model[0]} ({model[1]})")
            self.current_table2.setItem(i, 0, count_item)
            self.current_table2.setItem(i, 1, model_item)
            self.current_table2.verticalHeader().hide()
     
    def count_models(self):
        model_counts = {}

        for item in self.inventory:
            model_lang_key = f"{item['Modell']} {item['Sprache']}"

            if model_lang_key in model_counts:
                model_counts[model_lang_key] += 1
            else:
                model_counts[model_lang_key] = 1

        model_counts_str = '\n'.join([f"{count} {model_lang}" for model_lang, count in model_counts.items()])

        return model_counts_str
            
    def update_checkedout(self):
    # Leeren der Tabelle
        self.checkedout_table.setRowCount(0)

    # Einfügen der ausgebuchten Notebooks in die Tabelle
        for item in self.checked_out:
            row_position = self.checkedout_table.rowCount()
            self.checkedout_table.insertRow(row_position)

            self.checkedout_table.setItem(row_position, 0, QTableWidgetItem(item['Modell']))
            self.checkedout_table.setItem(row_position, 1, QTableWidgetItem(item['Sprache']))
            self.checkedout_table.setItem(row_position, 2, QTableWidgetItem(item['SN']))
            self.checkedout_table.setItem(row_position, 3, QTableWidgetItem(item.get('MAC-Adresse', '-')))
            self.checkedout_table.setItem(row_position, 4, QTableWidgetItem(item.get('Einbuchungsdatum', '-')))
            self.checkedout_table.setItem(row_position, 5, QTableWidgetItem(item.get('Ausbuchungsdatum', '-')))
            self.checkedout_table.setItem(row_position, 6, QTableWidgetItem(item.get('User', '-')))

    # Funktion zum Speichern der Inventardaten in die Datei
    def save_inventory(self):
        with open(INVENTORY_FILE, 'w') as f:
            portalocker.lock(f, portalocker.LOCK_EX)
            json.dump(self.inventory, f)
            portalocker.unlock(f)
        self.save_checkedout()

    def save_checkedout(self):
        with open(CHECKEDOUT_FILE, 'w') as f:
            portalocker.lock(f, portalocker.LOCK_EX)
            json.dump(self.checked_out, f)
            portalocker.unlock(f)
            
    def update_model_count(self):
        model_counts = self.count_models()

        # Entfernen aller Widgets aus dem Layout
        for i in reversed(range(self.model_count_layout.count())):
            self.model_count_layout.itemAt(i).widget().setParent(None)

        # Hinzufügen der Modelle und Sprachen zur Anzeige
        col = 0
        row = 0
        for model_lang, count in model_counts.items():
            label = QLabel(f"{count} {model_lang}")
            self.model_count_layout.addWidget(label, row, col)
            row += 1
            if row >= (len(model_counts) + 1) // 2:
                row = 0
                col += 1

    def count_models(self):
        model_counts = {}

        for item in self.inventory:
            model_lang_key = f"{item['Modell']} {item['Sprache']}"

            if model_lang_key in model_counts:
                model_counts[model_lang_key] += 1
            else:
                model_counts[model_lang_key] = 1

        return model_counts

    def load_data(self):
        with open(INVENTORY_FILE, 'r') as f:
            portalocker.lock(f, portalocker.LOCK_SH)
            self.inventory = json.load(f)
            portalocker.unlock(f)

        with open(CHECKEDOUT_FILE, 'r') as f:
            portalocker.lock(f, portalocker.LOCK_SH)
            self.checked_out = json.load(f)
            portalocker.unlock(f)

        self.update_inventory()
        self.update_checkedout()
   
    def change_identifier(self):
        identifier, ok = QInputDialog.getItem(self, "Identifier ändern", "Identifier auswählen:", ["Seriennummer", "MAC-Adresse", "Modell", "Sprache"])
        if ok and identifier:
            if identifier == "Seriennummer":
                self.change_serial_number()
            elif identifier == "MAC-Adresse":
                self.change_mac_address()
            elif identifier == "Modell":
                self.change_model()
            elif identifier == "Sprache":
                self.change_language()

    def change_serial_number(self):
        sn, ok = QtWidgets.QInputDialog.getText(self, "Seriennummer ändern", "Seriennummer eingeben:")
        if ok and sn:
            new_sn, ok = QtWidgets.QInputDialog.getText(self, "Seriennummer ändern", f"Neue Seriennummer für {sn}:")
            if ok and new_sn:
                for item in self.inventory:
                    if item["SN"] == sn:
                        item["SN"] = new_sn
                        self.save_inventory()
                        self.update_inventory()
                        break
                else:
                    QMessageBox.warning(self, "Gerät nicht gefunden", f"Es wurde kein Gerät mit der Seriennummer {sn} gefunden.")

    def change_mac_address(self):
        sn, ok = QtWidgets.QInputDialog.getText(self, "MAC-Adresse ändern", "Seriennummer eingeben:")
        if ok and sn:
            new_mac, ok = QtWidgets.QInputDialog.getText(self, "MAC-Adresse ändern", f"Neue MAC-Adresse für {sn}:")
            if ok and new_mac:
                new_mac = new_mac.replace(':', '')  # Entfernt Doppelpunkte aus der neuen MAC-Adresse
                for item in self.inventory:
                    if item["SN"] == sn:
                        item["MAC-Adresse"] = new_mac
                        self.save_inventory()
                        self.update_inventory()
                        break
                else:
                    QMessageBox.warning(self, "Gerät nicht gefunden", f"Es wurde kein Gerät mit der Seriennummer {sn} gefunden.")


    def change_model(self):
        sn, ok = QtWidgets.QInputDialog.getText(self, "Modell ändern", "Seriennummer eingeben:")
        if ok and sn:
            new_model, ok = QtWidgets.QInputDialog.getItem(self, "Modell ändern", f"Neues Modell für {sn} auswählen:", self.MODELS)
            if ok and new_model:
                for item in self.inventory:
                    if item["SN"] == sn:
                        item["Modell"] = new_model
                        self.save_inventory()
                        self.update_inventory()
                        break
                else:
                    QMessageBox.warning(self, "Gerät nicht gefunden", f"Es wurde kein Gerät mit der Seriennummer {sn} gefunden.")

    def change_language(self):
        sn, ok = QtWidgets.QInputDialog.getText(self, "Sprache ändern", "Seriennummer eingeben:")
        if ok and sn:
            new_language, ok = QtWidgets.QInputDialog.getItem(self, "Sprache ändern", f"Neue Sprache für {sn} auswählen:", self.LANGUAGES)
            if ok and new_language:
                for item in self.inventory:
                    if item["SN"] == sn:
                        item["Sprache"] = new_language
                        self.save_inventory()
                        self.update_inventory()
                        break
                else:
                    QMessageBox.warning(self, "Gerät nicht gefunden", f"Es wurde kein Gerät mit der Seriennummer {sn} gefunden.")





    def keyPressEvent(self, event):
        if event.modifiers() == Qt.ControlModifier:
            if event.key() == Qt.Key_1:
                # Timer für die Aktualisierung der Daten anhalten
                self.update_timer.stop()

                # Identifier-Änderungsfunktion aufrufen
                self.change_identifier()

                # Timer für die Aktualisierung der Daten wieder starten
                self.update_timer.start(2000)
            elif event.key() == Qt.Key_2:
                # Aufruf der Funktion zum Hinzufügen/Entfernen von Modellen
                self.modify_models()

    def modify_models(self):
        # Dialog zum Hinzufügen/Entfernen von Modellen
        self.setEnabled(False)
        action, ok = QtWidgets.QInputDialog.getItem(self, "Modelle ändern", "Aktion auswählen:", ["Modell hinzufügen", "Modell entfernen", "Modell umbenennen"])
        if ok and action:
            if action == "Modell hinzufügen":
                model, ok = QtWidgets.QInputDialog.getText(self, "Modelle ändern", "Modell eingeben:")
                if ok and model:
                    # Wenn das Modell nicht existiert, fügen Sie es hinzu
                    if model not in self.MODELS:
                        self.MODELS.append(model)
            elif action == "Modell entfernen":
                model, ok = QtWidgets.QInputDialog.getItem(self, "Modelle ändern", "Modell auswählen:", self.MODELS)
                if ok and model:
                    # Wenn das Modell bereits existiert, entfernen Sie es
                    if model in self.MODELS:
                        self.MODELS.remove(model)
            elif action == "Modell umbenennen":
                model, ok = QtWidgets.QInputDialog.getItem(self, "Modelle ändern", "Modell auswählen:", self.MODELS)
                if ok and model:
                    new_name, ok = QtWidgets.QInputDialog.getText(self, "Modelle ändern", "Neuen Namen eingeben:")
                    if ok and new_name:
                        # Finden Sie den Index des alten Modellnamens und ersetzen Sie ihn durch den neuen Namen
                        index = self.MODELS.index(model)
                        self.MODELS[index] = new_name
            # Speichern Sie die aktualisierte Modellliste
            with open('models.json', 'w') as f:
                json.dump(self.MODELS, f)
            # Aktualisieren Sie das Dropdown-Menü
            self.update_model_dropdown()
        self.setEnabled(True)




    def update_model_dropdown(self):
        # Leeren Sie das Dropdown-Menü
        self.checkin_model.clear()
        # Fügen Sie jedes Modell im Modell-Array zum Dropdown-Menü hinzu
        for model in self.MODELS:
            self.checkin_model.addItem(model)

if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setStyle(QStyleFactory.create("Fusion"))
    window = MainWindow()
    window.showMaximized()
    sys.exit(app.exec_())
